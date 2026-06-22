"""Generischer Bewertungs-Import aus XLSX-Dateien.

Erwartet eine Header-Zeile mit "ID" in Spalte A. Nach der Header-Zeile:
    A: anforderung_id
    F (col 5): bewertung (0-5)
    G (col 6): kommentar
    H (col 7): massnahme
    I (col 8): verantwortlich
    J (col 9): zieldatum (YYYY-MM-DD)

Wird von AI-Act und ähnlichen Modulen genutzt, die kein eigenes
io_xlsx-Modul haben. Modul-spezifische Importer (CRA, NIS2, DSGVO, Gutachten)
bleiben weiterhin in Modul-Code.
"""

from __future__ import annotations
import re
from pathlib import Path
from typing import Any, Iterable


def import_bewertungen(
    xlsx_path: Path,
    *,
    known_ids: Iterable[str] | None = None,
    expected_label: str = 'Fragebogen',
) -> list[dict[str, Any]]:
    """Liest Bewertungen aus einer XLSX-Datei.

    `known_ids` ist optional. Wenn gesetzt, werden Zeilen mit unbekannten IDs
    übersprungen.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError('openpyxl ist nicht installiert.') from exc

    try:
        from security_utils import validate_office_archive
        validate_office_archive(Path(xlsx_path), expected_suffix='.xlsx')
    except Exception:
        pass

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active

    MAX_ROWS = 5000
    if ws.max_row > MAX_ROWS:
        raise ValueError(f'XLSX hat zu viele Zeilen ({ws.max_row} > {MAX_ROWS})')

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=20):
        if str(row[0].value or '').strip().upper() == 'ID':
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError(f"Header-Zeile mit 'ID' nicht gefunden – ist dies ein {expected_label}?")

    ids_filter: set[str] | None = set(known_ids) if known_ids is not None else None
    results: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=header_row + 1):
        id_val = str(row[0].value or '').strip()
        if not id_val:
            continue
        if ids_filter is not None and id_val not in ids_filter:
            continue

        def _cell(col: int) -> str:
            return str(row[col].value or '').strip() if col < len(row) else ''

        try:
            bew_raw = _cell(5)
            bewertung = int(float(bew_raw)) if bew_raw else 0
        except (ValueError, TypeError):
            bewertung = 0
        bewertung = max(0, min(5, bewertung))

        zieldatum = _cell(9)
        if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', zieldatum):
            zieldatum = ''

        results.append({
            'anforderung_id': id_val,
            'bewertung': bewertung,
            'kommentar': _cell(6)[:4000],
            'massnahme': _cell(7)[:2000],
            'verantwortlich': _cell(8)[:200],
            'zieldatum': zieldatum,
        })

    return results
