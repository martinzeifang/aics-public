from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from security_utils import (
    MAX_XLSX_COLUMNS,
    MAX_XLSX_ROWS,
    safe_generated_file,
    validate_office_archive,
    workspace_root_from,
)
from shared.audit import audit_event
from shared.fs_perms import ensure_private_dir, ensure_private_file


@dataclass(frozen=True)
class XlsxItem:
    file_name: str
    sheet_name: str
    row: int
    layout: str  # "system" or "service"

    # Common
    title: str
    question: str

    # System layout fields
    schutzziel: str | None = None
    umsetzung: str | None = None
    bemerkung_umsetzung: str | None = None

    # Service layout fields
    baso_id: str | None = None
    contract_assured: str | None = None
    ops_met: str | None = None
    bemerkung: str | None = None


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def detect_layout(xlsx_path: Path) -> str:
    validate_office_archive(xlsx_path, expected_suffix=".xlsx")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws: Any = wb.active
    if ws is None:
        raise ValueError(f"No active worksheet in: {xlsx_path}")
    if ws.max_row > MAX_XLSX_ROWS or ws.max_column > MAX_XLSX_COLUMNS:
        raise ValueError(f"Workbook dimensions exceed safe limits: {xlsx_path.name}")

    for r in range(1, min(120, ws.max_row) + 1):
        row = [_cell_str(ws.cell(r, c).value) for c in range(1, min(30, ws.max_column) + 1)]
        joined = " | ".join(row)
        if "Herkunft der" in joined and "Zuordnung" in joined and "Bemerkung zur Umsetzung" in joined:
            return "system"
        if "BASO-ID" in row and "Vertraglich zugesichert?" in joined:
            return "service"

    first = _cell_str(ws.cell(1, 1).value)
    if "Anhang" in first and "Informationssicherheit" in first:
        return "service"

    raise ValueError(f"Unbekanntes XLSX-Layout: {xlsx_path}")


def read_items(xlsx_path: Path) -> list[XlsxItem]:
    layout = detect_layout(xlsx_path)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=False)
    ws: Any = wb.active
    if ws is None:
        raise ValueError(f"No active worksheet in: {xlsx_path}")
    if ws.max_row > MAX_XLSX_ROWS or ws.max_column > MAX_XLSX_COLUMNS:
        raise ValueError(f"Workbook dimensions exceed safe limits: {xlsx_path.name}")
    if layout == "system":
        return _read_system_items(xlsx_path, ws)
    if layout == "service":
        return _read_service_items(xlsx_path, ws)
    raise AssertionError("unreachable")


def _find_header_row(ws: Any, required: set[str], scan_rows: int = 160) -> tuple[int, dict[str, int]]:
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        values = {c: _cell_str(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        inv = {v: c for c, v in values.items() if v}
        if required.issubset(inv.keys()):
            return r, {k: inv[k] for k in required}
    raise ValueError("Header row not found")


def _read_system_items(xlsx_path: Path, ws: Any) -> list[XlsxItem]:
    required = {
        "Bezeichnung",
        "Inhalt / Anforderung",
        "Zuordnung zu Schutzzielen",
        "Umsetzung",
        "Bemerkung zur Umsetzung",
    }
    header_row, cols = _find_header_row(ws, required)

    out: list[XlsxItem] = []
    for r in range(header_row + 1, ws.max_row + 1):
        title = _cell_str(ws.cell(r, cols["Bezeichnung"]).value)
        question = _cell_str(ws.cell(r, cols["Inhalt / Anforderung"]).value)
        if not title and not question:
            continue
        schutzziel = _cell_str(ws.cell(r, cols["Zuordnung zu Schutzzielen"]).value) or None
        umsetzung = _cell_str(ws.cell(r, cols["Umsetzung"]).value) or None
        bemerkung = _cell_str(ws.cell(r, cols["Bemerkung zur Umsetzung"]).value) or None
        out.append(
            XlsxItem(
                file_name=xlsx_path.name,
                sheet_name=ws.title,
                row=r,
                layout="system",
                title=title,
                question=question,
                schutzziel=schutzziel,
                umsetzung=umsetzung,
                bemerkung_umsetzung=bemerkung,
            )
        )
    return out


def _read_service_items(xlsx_path: Path, ws: Any) -> list[XlsxItem]:
    header_row = None
    for r in range(1, min(200, ws.max_row) + 1):
        row = [_cell_str(ws.cell(r, c).value) for c in range(1, min(20, ws.max_column) + 1)]
        if "BASO-ID" in row and "Bemerkung zur Umsetzung" in row:
            header_row = r
            break
    if header_row is None:
        raise ValueError("Service header row not found")

    # Known template columns (1-based):
    # 1 BASO-ID
    # 4 Sollmassnahmen-Bezeichnung
    # 5 Sollmassnahme
    # 8 Vertraglich zugesichert?
    # 11 Umsetzung ... erfuellt?
    # 12 Bemerkung zur Umsetzung
    c_baso = 1
    c_title = 4
    c_question = 5
    c_contract = 8
    c_ops = 11
    c_bem = 12

    out: list[XlsxItem] = []
    for r in range(header_row + 1, ws.max_row + 1):
        baso_id = _cell_str(ws.cell(r, c_baso).value)
        title = _cell_str(ws.cell(r, c_title).value)
        question = _cell_str(ws.cell(r, c_question).value)
        if not baso_id and not title and not question:
            continue
        contract = _cell_str(ws.cell(r, c_contract).value) or None
        ops = _cell_str(ws.cell(r, c_ops).value) or None
        bem = _cell_str(ws.cell(r, c_bem).value) or None
        out.append(
            XlsxItem(
                file_name=xlsx_path.name,
                sheet_name=ws.title,
                row=r,
                layout="service",
                title=title,
                question=question,
                baso_id=baso_id or None,
                contract_assured=contract,
                ops_met=ops,
                bemerkung=bem,
            )
        )
    return out


def write_answers_to_copy(src_xlsx: Path, dst_xlsx: Path, updates: list[dict[str, Any]]) -> None:
    validate_office_archive(src_xlsx, expected_suffix=".xlsx")
    wb = openpyxl.load_workbook(str(src_xlsx))
    ws: Any = wb[wb.sheetnames[0]]
    if ws.max_row > MAX_XLSX_ROWS or ws.max_column > MAX_XLSX_COLUMNS:
        raise ValueError(f"Workbook dimensions exceed safe limits: {src_xlsx.name}")
    layout = detect_layout(src_xlsx)

    if layout == "system":
        required = {
            "Bezeichnung",
            "Inhalt / Anforderung",
            "Zuordnung zu Schutzzielen",
            "Umsetzung",
            "Bemerkung zur Umsetzung",
        }
        _header_row, cols = _find_header_row(ws, required)
        for u in updates:
            r = int(u["row"])
            if u.get("umsetzung") is not None:
                ws.cell(r, cols["Umsetzung"]).value = u["umsetzung"]
            if u.get("schutzziel") is not None:
                ws.cell(r, cols["Zuordnung zu Schutzzielen"]).value = u["schutzziel"]
            if u.get("bemerkung_umsetzung") is not None:
                ws.cell(r, cols["Bemerkung zur Umsetzung"]).value = u["bemerkung_umsetzung"]

    elif layout == "service":
        # Known template columns (1-based)
        c_contract_by = 6
        c_contract_date = 7
        c_contract_assured = 8
        c_ops_by = 9
        c_ops_date = 10
        c_ops_met = 11
        c_bem = 12

        for u in updates:
            r = int(u["row"])
            if u.get("contract_assured") is not None:
                ws.cell(r, c_contract_assured).value = u["contract_assured"]
            if u.get("ops_met") is not None:
                ws.cell(r, c_ops_met).value = u["ops_met"]
            if u.get("bemerkung") is not None:
                ws.cell(r, c_bem).value = u["bemerkung"]
            if u.get("evaluated_by") is not None:
                ws.cell(r, c_contract_by).value = u["evaluated_by"]
                ws.cell(r, c_ops_by).value = u["evaluated_by"]
            if u.get("evaluated_at") is not None:
                ws.cell(r, c_contract_date).value = u["evaluated_at"]
                ws.cell(r, c_ops_date).value = u["evaluated_at"]

    else:
        raise AssertionError("unreachable")

    dst_xlsx = safe_generated_file(dst_xlsx, workspace_root_from(Path(__file__)))
    ensure_private_dir(dst_xlsx.parent)
    wb.save(str(dst_xlsx))
    ensure_private_file(dst_xlsx)
    audit_event("export.write", module="baso", outcome="success", details={"path": str(dst_xlsx), "kind": "xlsx"})
