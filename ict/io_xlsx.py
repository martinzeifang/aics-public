from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from security_utils import (
    MAX_XLSX_COLUMNS,
    MAX_XLSX_ROWS,
    safe_generated_file,
    validate_office_archive,
    workspace_root_from,
)
from shared.audit import audit_event
from shared.fs_perms import ensure_private_dir, ensure_private_file
from shared.xlsx_safety import safe_cell_value


@dataclass(frozen=True)
class IctItem:
    file_name: str
    sheet_name: str
    row: int
    question_id: str
    title: str
    question: str
    answer: str | None = None
    maturity: int | None = None
    explanation: str | None = None
    guidance: str | None = None
    optimization_potential: str | None = None


def make_item_id(it: IctItem) -> str:
    return f"{it.file_name}::{it.sheet_name}::R{it.row}"


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def detect_layout(xlsx_path: Path) -> str:
    validate_office_archive(xlsx_path, expected_suffix=".xlsx")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws: Any = wb.active
    if ws is None:
        raise ValueError(f"No active worksheet in: {xlsx_path}")
    if ws.max_row > MAX_XLSX_ROWS or ws.max_column > MAX_XLSX_COLUMNS:
        raise ValueError(f"Workbook dimensions exceed safe limits: {xlsx_path.name}")

    for r in range(1, min(80, ws.max_row) + 1):
        row = [_cell_str(ws.cell(r, c).value) for c in range(1, min(12, ws.max_column) + 1)]
        joined = " | ".join(row)
        if (
            "Reifegrad" in joined
            and "Nr." in joined
            and (
                "Fragen zu den IT-Risikokontrollen" in joined
                or "Fragen zu IT-Risikokontrollen" in joined
                or "Fragen" in joined
            )
        ):
            return "ict"
    raise ValueError(f"Unbekanntes ICT-XLSX-Layout: {xlsx_path}")


def read_items(xlsx_path: Path) -> list[IctItem]:
    detect_layout(xlsx_path)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=False)
    ws: Any = wb.active
    if ws is None:
        raise ValueError(f"No active worksheet in: {xlsx_path}")
    if ws.max_row > MAX_XLSX_ROWS or ws.max_column > MAX_XLSX_COLUMNS:
        raise ValueError(f"Workbook dimensions exceed safe limits: {xlsx_path.name}")
    return _read_ict_items(xlsx_path, ws)


def _is_instruction(text: str) -> bool:
    s = (text or "").strip().casefold()
    return s.startswith("begr") or "selbsteinsch" in s or "reifegrad" in s


def _read_ict_items(xlsx_path: Path, ws: Any) -> list[IctItem]:
    c_id = 2
    c_question = 3
    c_answer = 4
    c_maturity = 5
    c_explanation = 6
    c_extra = 7

    headings: list[str] = []
    out: list[IctItem] = []
    for r in range(1, ws.max_row + 1):
        raw_id = _cell_str(ws.cell(r, c_id).value)
        question = _cell_str(ws.cell(r, c_question).value)
        answer = _cell_str(ws.cell(r, c_answer).value)
        maturity_raw = _cell_str(ws.cell(r, c_maturity).value)
        explanation_cell = _cell_str(ws.cell(r, c_explanation).value)

        if raw_id.startswith("RK"):
            guidance = explanation_cell if _is_instruction(explanation_cell) else ""
            explanation = ""
            if guidance:
                next_id = _cell_str(ws.cell(r + 1, c_id).value) if r < ws.max_row else ""
                next_expl = _cell_str(ws.cell(r + 1, c_explanation).value) if r < ws.max_row else ""
                next_opt = _cell_str(ws.cell(r + 1, c_extra).value) if r < ws.max_row else ""
                if not next_id and next_expl and not _is_instruction(next_expl) and next_expl.casefold() != "erläuterungen (inkl. stärken)".casefold():
                    explanation = next_expl
                    optimization = next_opt
                else:
                    optimization = ""
            else:
                explanation = explanation_cell
                optimization = _cell_str(ws.cell(r, c_extra).value)

            maturity = None
            if maturity_raw:
                try:
                    maturity = int(float(maturity_raw))
                except Exception:
                    maturity = None

            out.append(
                IctItem(
                    file_name=xlsx_path.name,
                    sheet_name=ws.title,
                    row=r,
                    question_id=raw_id,
                    title=" / ".join(headings[-2:]),
                    question=question,
                    answer=answer or None,
                    maturity=maturity,
                    explanation=explanation or None,
                    guidance=guidance or None,
                    optimization_potential=optimization or None,
                )
            )
            continue

        if question and not answer and not maturity_raw and not explanation_cell and not _cell_str(ws.cell(r, c_extra).value):
            headings.append(question)


    return out


def needs_answer(it: IctItem) -> bool:
    if not (it.answer or "").strip():
        return True
    if it.maturity is None:
        return True
    if not (it.explanation or "").strip():
        return True
    return False


def _explanation_target_row(ws: Any, row: int) -> int:
    current = _cell_str(ws.cell(row, 6).value)
    if _is_instruction(current):
        next_id = _cell_str(ws.cell(row + 1, 2).value) if row < ws.max_row else ""
        if not next_id and row < ws.max_row:
            return row + 1
    return row


def write_answers_to_copy(src_xlsx: Path, dst_xlsx: Path, updates: list[dict[str, Any]]) -> None:
    detect_layout(src_xlsx)
    wb = openpyxl.load_workbook(str(src_xlsx))
    ws: Any = wb[wb.sheetnames[0]]
    if ws.max_row > MAX_XLSX_ROWS or ws.max_column > MAX_XLSX_COLUMNS:
        raise ValueError(f"Workbook dimensions exceed safe limits: {src_xlsx.name}")

    for u in updates:
        row = int(u["row"])
        if u.get("answer") is not None:
            ws.cell(row, 4).value = safe_cell_value(u["answer"])
        if u.get("maturity") is not None:
            ws.cell(row, 5).value = int(u["maturity"])
        if u.get("explanation") is not None:
            target_row = _explanation_target_row(ws, row)
            ws.cell(target_row, 6).value = safe_cell_value(u["explanation"])
        if u.get("optimization_potential") is not None:
            target_row = _explanation_target_row(ws, row)
            ws.cell(target_row, 7).value = safe_cell_value(u["optimization_potential"])

    for it in _read_ict_items(src_xlsx, ws):
        maturity_value = ws.cell(it.row, 5).value
        try:
            maturity = int(float(maturity_value)) if maturity_value is not None and str(maturity_value).strip() else None
        except Exception:
            maturity = None
        if maturity == 1:
            target_row = _explanation_target_row(ws, it.row)
            ws.cell(target_row, 7).value = "nicht erforderlich"

    dst_xlsx = safe_generated_file(dst_xlsx, workspace_root_from(Path(__file__)))
    ensure_private_dir(dst_xlsx.parent)
    wb.save(str(dst_xlsx))
    ensure_private_file(dst_xlsx)
    audit_event("export.write", module="ict", outcome="success", details={"path": str(dst_xlsx), "kind": "xlsx"})
