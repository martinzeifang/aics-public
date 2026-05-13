"""Excel export and import for Risikobewertung."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from security_utils import safe_generated_file, workspace_root_from
from shared.audit import audit_event
from shared.fs_perms import ensure_private_dir, ensure_private_file

from risikobewertung.frameworks import (
    FRAMEWORK_LABELS,
    berechne_risiko,
    framework_felder,
)

# ── Column palette ────────────────────────────────────────────────────────────
_COL_HEADER_BG = "003078"
_COL_HEADER_FG = "FFFFFF"
_COL_TITLE_BG  = "0060a8"
_COL_EDITABLE  = "FFFDE7"   # light yellow – user fills these in
_COL_COMPUTED  = "E3F2FD"   # light blue  – calculated fields
_COL_BORDER    = "B8D0E4"

_FW_BG = {
    "Finanzinstitute": "FCE4EC",
    "STRIDE":          "E8F0FB",
    "HEAVENS":         "E8F5E9",
    "OCTAVE":          "FFF8E1",
    "TARA":            "F3E5F5",
}

# Fixed columns (always present)
_COL_NR         = 1   # A
_COL_NAME       = 2   # B
_COL_BESCHR     = 3   # C
_COL_FRAMEWORK  = 4   # D
# Framework-specific columns start at E (5)
_DATA_START_COL = 5

_HEADER_ROW    = 2
_DATA_START_ROW = 3


# ── Export ────────────────────────────────────────────────────────────────────

def export_risiken(
    risks: list[dict[str, Any]],
    out_path: Path,
    projekt_name: str,
    framework: str,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError("openpyxl ist nicht installiert.") from exc

    felder_defs = framework_felder(framework)
    fw_label = FRAMEWORK_LABELS.get(framework, framework)
    fw_bg = _FW_BG.get(framework, "FFFFFF")

    wb = Workbook()
    ws = wb.active
    ws.title = "Risiken"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=_DATA_START_COL + len(felder_defs) + 3)
    title_cell = ws.cell(row=1, column=1,
                         value=f"Risikobewertung – {projekt_name} – Framework: {fw_label}")
    title_cell.font   = Font(bold=True, color="FFFFFF", size=12)
    title_cell.fill   = PatternFill("solid", fgColor=_COL_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    # Build header row
    thin = Side(style="thin", color=_COL_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(col: int, text: str) -> None:
        c = ws.cell(row=_HEADER_ROW, column=col, value=text)
        c.font      = Font(bold=True, color=_COL_HEADER_FG, size=9)
        c.fill      = PatternFill("solid", fgColor=_COL_HEADER_BG)
        c.border    = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    hdr(_COL_NR,        "Nr")
    hdr(_COL_NAME,      "Risikoname")
    hdr(_COL_BESCHR,    "Beschreibung")
    hdr(_COL_FRAMEWORK, "Framework")

    for i, fd in enumerate(felder_defs, start=_DATA_START_COL):
        hdr(i, fd["label"])

    # Score + label after framework fields
    score_col = _DATA_START_COL + len(felder_defs)
    label_col = score_col + 1
    detail_col = label_col + 1
    hdr(score_col,  "Risikowert")
    hdr(label_col,  "Risikolevel")
    hdr(detail_col, "Berechnung")

    ws.row_dimensions[_HEADER_ROW].height = 28
    ws.freeze_panes = ws.cell(row=_DATA_START_ROW, column=1)

    # Add data validation dropdowns for each combo field
    for i, fd in enumerate(felder_defs, start=_DATA_START_COL):
        if fd["typ"] == "combo" and fd["optionen"]:
            opts = ",".join(f'"{o}"' for o in fd["optionen"])
            dv = DataValidation(type="list", formula1=opts, allow_blank=True)
            dv.sqref = (
                f"{get_column_letter(i)}{_DATA_START_ROW}:"
                f"{get_column_letter(i)}{_DATA_START_ROW + max(len(risks), 200)}"
            )
            ws.add_data_validation(dv)

    # Data rows
    for r in risks:
        row_idx = _DATA_START_ROW + (r.get("nr", 1) - 1)
        felder = r.get("felder", {})

        def dcell(col: int, val, editable: bool = False, computed: bool = False) -> None:
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border    = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
            bg = _COL_EDITABLE if editable else (_COL_COMPUTED if computed else "FFFFFF")
            c.fill = PatternFill("solid", fgColor=bg)

        dcell(_COL_NR,        r.get("nr", ""))
        dcell(_COL_NAME,      r.get("risk_name", ""), editable=True)
        dcell(_COL_BESCHR,    r.get("beschreibung", ""), editable=True)
        dcell(_COL_FRAMEWORK, fw_label)

        for i, fd in enumerate(felder_defs, start=_DATA_START_COL):
            dcell(i, felder.get(fd["key"], ""), editable=True)

        dcell(score_col,  r.get("risikowert"), computed=True)
        dcell(label_col,  r.get("risiko_label", ""), computed=True)
        dcell(detail_col, r.get("detail_text", ""), computed=True)

        ws.row_dimensions[row_idx].height = 50

    # Column widths
    ws.column_dimensions[get_column_letter(_COL_NR)].width = 5
    ws.column_dimensions[get_column_letter(_COL_NAME)].width = 28
    ws.column_dimensions[get_column_letter(_COL_BESCHR)].width = 40
    ws.column_dimensions[get_column_letter(_COL_FRAMEWORK)].width = 20
    for i in range(_DATA_START_COL, score_col):
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.column_dimensions[get_column_letter(score_col)].width = 12
    ws.column_dimensions[get_column_letter(label_col)].width = 18
    ws.column_dimensions[get_column_letter(detail_col)].width = 35

    out_path = safe_generated_file(out_path, workspace_root_from(Path(__file__)))
    ensure_private_dir(out_path.parent)
    wb.save(str(out_path))
    ensure_private_file(out_path)
    audit_event("export.write", module="risikobewertung", outcome="success", details={"path": str(out_path), "kind": "xlsx"})


# ── Import ────────────────────────────────────────────────────────────────────

def import_risiken(xlsx_path: Path, framework: str) -> list[dict[str, Any]]:
    """Read risks from an exported Excel file.

    Accepts files exported by export_risiken() as well as simplified files with
    columns: Nr | Risikoname | Beschreibung | <felder...>
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl ist nicht installiert.") from exc

    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active

    felder_defs = framework_felder(framework)
    fd_labels   = [fd["label"] for fd in felder_defs]
    fd_keys     = [fd["key"]   for fd in felder_defs]

    # Find header row: first row whose cell A contains a numeric value OR
    # cell B contains "Risikoname"
    header_row = None
    col_map: dict[str, int] = {}  # label -> col index (1-based)
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
        vals = [str(c.value or "").strip() for c in row]
        if "Risikoname" in vals or "risikoname" in [v.lower() for v in vals]:
            header_row = ridx
            for cidx, v in enumerate(vals, start=1):
                col_map[v] = cidx
                col_map[v.lower()] = cidx
            break

    data_start = (header_row or _HEADER_ROW) + 1

    results: list[dict] = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if all(v is None for v in row):
            continue

        def _get(label: str, fallback_col: int) -> str:
            col = col_map.get(label) or col_map.get(label.lower()) or fallback_col
            raw = row[col - 1] if col <= len(row) else None
            return str(raw).strip() if raw is not None else ""

        name = _get("Risikoname", _COL_NAME)
        if not name:
            continue

        beschr = _get("Beschreibung", _COL_BESCHR)
        felder: dict[str, str] = {}
        for fd_label, fd_key in zip(fd_labels, fd_keys):
            felder[fd_key] = _get(fd_label, 0)

        # Re-calculate score from imported values
        score, label, detail = berechne_risiko(framework, felder)

        results.append({
            "risk_name":    name,
            "beschreibung": beschr,
            "framework":    framework,
            "felder":       felder,
            "risikowert":   score,
            "risiko_label": label,
            "detail_text":  detail,
            "bewertung_text": "",
            "prompt_text":    "",
        })

    return results
