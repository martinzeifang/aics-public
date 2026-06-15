"""Excel-Import und -Export für Gutachten-Fragebögen.

Spaltenlayout des generierten Fragebogens:
    A: Nr
    B: Framework
    C: Kapitel/Artikel
    D: Thema
    E: Interviewfrage
    F: Antwort          ← vom Interviewten auszufüllen
    G: Bewertung        ← Dropdown: erfüllt / teilweise erfüllt / nicht erfüllt / nicht anwendbar
    H: Kommentar
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from security_utils import safe_generated_file, workspace_root_from
from shared.audit import audit_event
from shared.fs_perms import ensure_private_dir, ensure_private_file
from shared.xlsx_safety import safe_cell_value
from security_utils import validate_office_archive

# ── Spalten-Mapping ───────────────────────────────────────────────────────────
COL_NR = 1
COL_FRAMEWORK = 2
COL_REF = 3
COL_THEMA = 4
COL_FRAGE = 5
COL_ANTWORT = 6
COL_BEWERTUNG = 7
COL_KOMMENTAR = 8

HEADER_ROW = 2          # Zeile 1 = Titelzeile, Zeile 2 = Spaltenköpfe
DATA_START_ROW = 3

BEWERTUNG_WERTE = ["erfüllt", "teilweise erfüllt", "nicht erfüllt", "nicht anwendbar"]

# Farben
COLOR_HEADER_BG = "003078"      # Dunkel-Navy (Logofarbe)
COLOR_HEADER_FG = "FFFFFF"
COLOR_TITLE_BG  = "0060a8"      # Logo-Blau
COLOR_FRAMEWORK_BG = {
    "DORA":     "E8F0FB",
    "NIS2":     "E8F5E9",
    "CRA":      "FFF8E1",
    "ISO27001": "F3E5F5",
}
COLOR_ANSWER_BG = "FFFDE7"      # Antwortfelder leicht gelb
COLOR_BORDER    = "B8D0E4"


def _thin_border() -> Border:
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


# ── Export ────────────────────────────────────────────────────────────────────

def export_fragebogen(
    questions: list[dict[str, Any]],
    out_path: Path,
    projekt_name: str,
    frameworks: list[str],
    bewertung_skala: list[str] | None = None,
) -> None:
    """Exportiert Interviewfragen als formatiertes Excel."""
    bewertung_skala = bewertung_skala or BEWERTUNG_WERTE

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fragebogen"

    # Freeze panes unter Kopfzeilen
    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=1)

    # ── Titelzeile ──
    ts = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = safe_cell_value(f"Compliance-Fragebogen – {projekt_name} – {ts}")
    title_cell.font = Font(bold=True, size=13, color=COLOR_HEADER_FG, name="Segoe UI")
    title_cell.fill = _make_fill(COLOR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[1].height = 28

    # ── Kopfzeile ──
    headers = ["Nr", "Framework", "Kapitel/Artikel", "Thema",
               "Interviewfrage", "Antwort", "Bewertung", "Kommentar"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=col, value=h)
        c.font = Font(bold=True, color=COLOR_HEADER_FG, name="Segoe UI", size=10)
        c.fill = _make_fill(COLOR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin_border()
    ws.row_dimensions[HEADER_ROW].height = 20

    # ── Daten-Validierung für Bewertung ──
    dv_formula = '"' + ",".join(bewertung_skala) + '"'
    dv = DataValidation(
        type="list",
        formula1=dv_formula,
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"G{DATA_START_ROW}:G{DATA_START_ROW + len(questions) + 100}"
    ws.add_data_validation(dv)

    # ── Datenzellen ──
    for row_idx, q in enumerate(questions, start=DATA_START_ROW):
        fw = str(q.get("framework", ""))
        row_color = COLOR_FRAMEWORK_BG.get(fw, "FFFFFF")

        def _cell(col: int, value: Any, wrap: bool = False, answer: bool = False) -> None:
            c = ws.cell(row=row_idx, column=col, value=safe_cell_value(value))
            c.font = Font(name="Segoe UI", size=10)
            c.fill = _make_fill(COLOR_ANSWER_BG if answer else row_color)
            c.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=wrap
            )
            c.border = _thin_border()

        _cell(COL_NR, q.get("question_num", row_idx - DATA_START_ROW + 1))
        _cell(COL_FRAMEWORK, fw)
        _cell(COL_REF, str(q.get("section_ref", "")))
        _cell(COL_THEMA, str(q.get("thema", "")))
        _cell(COL_FRAGE, str(q.get("frage", "")), wrap=True)
        _cell(COL_ANTWORT, str(q.get("antwort", "")), wrap=True, answer=True)
        _cell(COL_BEWERTUNG, str(q.get("bewertung", "")), answer=True)
        _cell(COL_KOMMENTAR, str(q.get("kommentar", "")), wrap=True, answer=True)

        ws.row_dimensions[row_idx].height = 60

    # ── Spaltenbreiten ──
    col_widths = {
        COL_NR: 6,
        COL_FRAMEWORK: 12,
        COL_REF: 18,
        COL_THEMA: 22,
        COL_FRAGE: 55,
        COL_ANTWORT: 45,
        COL_BEWERTUNG: 20,
        COL_KOMMENTAR: 35,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Metadaten-Blatt ──
    ws_meta = wb.create_sheet("Metadaten")
    ws_meta.column_dimensions["A"].width = 24
    ws_meta.column_dimensions["B"].width = 50
    meta_rows = [
        ("Projektname", projekt_name),
        ("Frameworks", ", ".join(frameworks)),
        ("Erstellt am", ts),
        ("Anzahl Fragen", len(questions)),
        ("Bewertungsskala", ", ".join(bewertung_skala)),
    ]
    for r, (k, v) in enumerate(meta_rows, start=1):
        ws_meta.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws_meta.cell(row=r, column=2, value=str(v))

    out_path = safe_generated_file(out_path, workspace_root_from(Path(__file__)))
    ensure_private_dir(out_path.parent)
    wb.save(str(out_path))
    ensure_private_file(out_path)
    audit_event("export.write", module="gutachten", outcome="success", details={"path": str(out_path), "kind": "xlsx"})


def fragebogen_filename(projekt_name: str) -> str:
    """Erstellt den Dateinamen mit Projektname, Datum und Uhrzeit."""
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in projekt_name)
    return f"Fragebogen_{safe_name}_{ts}.xlsx"


# ── Import ────────────────────────────────────────────────────────────────────

@dataclass
class ImportedQuestion:
    question_num: int
    framework: str
    section_ref: str
    thema: str
    frage: str
    antwort: str
    bewertung: str
    kommentar: str


def import_fragebogen(xlsx_path: Path) -> tuple[str, list[ImportedQuestion]]:
    """Liest einen ausgefüllten Fragebogen ein.

    Gibt (projekt_name, list[ImportedQuestion]) zurück.
    projekt_name wird aus dem Metadaten-Blatt gelesen, fällt zurück auf Dateinamen.
    """
    xlsx_path = Path(xlsx_path)
    validate_office_archive(xlsx_path, expected_suffix=".xlsx")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    # Projektname aus Metadaten-Blatt
    projekt_name = xlsx_path.stem
    if "Metadaten" in wb.sheetnames:
        ws_meta = wb["Metadaten"]
        for row in ws_meta.iter_rows(min_row=1, max_row=10, values_only=True):
            if row and str(row[0]).strip().lower() == "projektname" and row[1]:
                projekt_name = str(row[1]).strip()
                break

    ws = wb["Fragebogen"]
    questions: list[ImportedQuestion] = []

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            nr = int(row[COL_NR - 1]) if row[COL_NR - 1] is not None else 0
        except (ValueError, TypeError):
            continue

        def _str(v: Any) -> str:
            return str(v).strip() if v is not None else ""

        questions.append(ImportedQuestion(
            question_num=nr,
            framework=_str(row[COL_FRAMEWORK - 1]),
            section_ref=_str(row[COL_REF - 1]),
            thema=_str(row[COL_THEMA - 1]),
            frage=_str(row[COL_FRAGE - 1]),
            antwort=_str(row[COL_ANTWORT - 1]),
            bewertung=_str(row[COL_BEWERTUNG - 1]),
            kommentar=_str(row[COL_KOMMENTAR - 1]),
        ))

    wb.close()
    return projekt_name, questions
