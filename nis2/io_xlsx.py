"""NIS2-Modul – Excel Fragebogen Export und Import."""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from security_utils import safe_generated_file, workspace_root_from
from shared.audit import audit_event
from shared.fs_perms import ensure_private_dir, ensure_private_file
from shared.xlsx_safety import safe_cell_value

from nis2.requirements import (
    BEWERTUNG_LABELS,
    BEWERTUNG_SKALA,
    EINRICHTUNGSKLASSEN,
    KAPITEL,
    NIS2_ANFORDERUNGEN,
)

_NIS2_BLUE = "1A237E"
_KAPITEL_FARBEN = {
    "NIS1": {"header": "1565C0", "soft": "E3F2FD", "text": "FFFFFF"},
    "NIS2": {"header": "4A148C", "soft": "F3E5F5", "text": "FFFFFF"},
    "NIS3": {"header": "B71C1C", "soft": "FFEBEE", "text": "FFFFFF"},
    "NIS4": {"header": "E65100", "soft": "FFF3E0", "text": "FFFFFF"},
    "NIS5": {"header": "1B5E20", "soft": "E8F5E9", "text": "FFFFFF"},
}
_BEWERTUNG_FARBEN = {
    0: "9E9E9E", 1: "C62828", 2: "E65100",
    3: "F57F17", 4: "2E7D32", 5: "1B5E20",
}

_COLS = [
    ("ID",               8,  "A"),
    ("Anforderung",      42, "B"),
    ("Referenz (NIS2)",  28, "C"),
    ("Beschreibung",     52, "D"),
    ("Hinweise",         45, "E"),
    ("Bewertung (0-5)",  18, "F"),
    ("Kommentar / Nachweis", 40, "G"),
    ("Maßnahme",         35, "H"),
    ("Verantwortlich",   20, "I"),
    ("Zieldatum",        14, "J"),
]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, size: int = 10, color: str = "000000", italic: bool = False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")


def _border_thin() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(wrap: bool = True, valign: str = "top", halign: str = "left") -> Alignment:
    return Alignment(wrap_text=wrap, vertical=valign, horizontal=halign)


def _safe_filename(s: str) -> str:
    s = re.sub(r'[\\/:*?"<>|]', "-", s or "")
    s = re.sub(r"\s+", "_", s).strip("_")
    return s[:100] or "NIS2_Fragebogen"


def export_fragebogen(
    *,
    out_dir: Path,
    projekt_name: str,
    unternehmen: str = "",
    einrichtungsklasse: str = "wesentlich",
    berater: str = "",
    bestehende_bewertungen: dict[str, dict[str, Any]] | None = None,
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"NIS2_Fragebogen_{_safe_filename(projekt_name)}_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NIS2-Fragebogen"

    bew = bestehende_bewertungen or {}
    ekl_label = EINRICHTUNGSKLASSEN.get(einrichtungsklasse, {}).get("label", einrichtungsklasse)

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "NIS2-Readiness Questionnaire"
    c.font = _font(bold=True, size=20, color="FFFFFF")
    c.fill = _fill(_NIS2_BLUE)
    c.alignment = _align(valign="center", halign="center")
    ws.row_dimensions[1].height = 46

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = "Richtlinie (EU) 2022/2555 – NIS2 Cybersicherheitsrichtlinie  |  AI Compliance Suite"
    c.font = _font(size=10, color="9FA8DA", italic=True)
    c.fill = _fill("0D1B6E")
    c.alignment = _align(valign="center", halign="center")

    meta = [
        ("Projekt:", projekt_name),
        ("Unternehmen:", unternehmen),
        ("Einrichtungsklasse:", ekl_label),
        ("Berater:", berater),
        ("Erstellt am:", datetime.now().strftime("%d.%m.%Y")),
        ("Rechtsgrundlage:", "Richtlinie (EU) 2022/2555 (NIS2)"),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=label).font = _font(bold=True, size=10)
        ws.cell(row=i, column=2, value=safe_cell_value(value)).font = _font(size=10)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=False)
        ws.row_dimensions[i].height = 16

    ws.merge_cells("F3:J3")
    ws["F3"].value = "BEWERTUNGSSKALA"
    ws["F3"].font = _font(bold=True, size=9, color="FFFFFF")
    ws["F3"].fill = _fill("37474F")
    ws["F3"].alignment = _align(halign="center", valign="center")

    for i, (bval, binfo) in enumerate(BEWERTUNG_SKALA.items()):
        row = 4 + i
        ws.cell(row=row, column=6, value=bval).font = _font(bold=True, size=9, color="FFFFFF")
        ws.cell(row=row, column=6).fill = _fill(_BEWERTUNG_FARBEN[bval])
        ws.cell(row=row, column=6).alignment = _align(halign="center", valign="center")
        ws.merge_cells(f"G{row}:J{row}")
        ws.cell(row=row, column=7, value=binfo["label"]).font = _font(size=9)
        ws.cell(row=row, column=7).fill = _fill("F5F5F5")

    for col_idx, (_, width, _) in enumerate(_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    header_row = 10
    ws.row_dimensions[header_row].height = 32
    for col_idx, (header, _, _) in enumerate(_COLS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _font(bold=True, size=10, color="FFFFFF")
        cell.fill = _fill(_NIS2_BLUE)
        cell.alignment = _align(halign="center", valign="center")
        cell.border = _border_thin()

    dv = DataValidation(
        type="list",
        formula1='"0,1,2,3,4,5"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Ungültige Bewertung",
        error="Bitte einen Wert zwischen 0 und 5 eingeben.",
    )
    ws.add_data_validation(dv)

    current_row = header_row + 1
    prev_kapitel = None

    for req in NIS2_ANFORDERUNGEN:
        kap_id = req["kapitel"]
        kap_info = KAPITEL[kap_id]
        kap_farben = _KAPITEL_FARBEN[kap_id]

        if kap_id != prev_kapitel:
            prev_kapitel = kap_id
            ws.merge_cells(f"A{current_row}:J{current_row}")
            kap_cell = ws.cell(
                row=current_row, column=1,
                value=f"  {kap_info['titel']}  ·  {kap_info['referenz']}",
            )
            kap_cell.font = _font(bold=True, size=11, color="FFFFFF")
            kap_cell.fill = _fill(kap_farben["header"])
            kap_cell.alignment = _align(valign="center", halign="left")
            ws.row_dimensions[current_row].height = 28
            current_row += 1

        rid = req["id"]
        bestehend = bew.get(rid, {})
        row_bg = kap_farben["soft"] if (current_row % 2 == 0) else "FFFFFF"

        values = [
            rid,
            req["titel"],
            req["ref"],
            req["beschreibung"],
            req["hinweise"],
            bestehend.get("bewertung", 0),
            bestehend.get("kommentar", ""),
            bestehend.get("massnahme", ""),
            bestehend.get("verantwortlich", ""),
            bestehend.get("zieldatum", ""),
        ]

        ws.row_dimensions[current_row].height = 60
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=safe_cell_value(value))
            cell.border = _border_thin()
            cell.alignment = _align()

            if col_idx == 1:
                cell.font = _font(bold=True, size=9, color=kap_farben["header"])
                cell.fill = _fill(kap_farben["soft"])
            elif col_idx == 2:
                cell.font = _font(bold=True, size=10)
                cell.fill = _fill(row_bg)
            elif col_idx == 6:
                bval = int(value) if value else 0
                cell.font = _font(bold=True, size=11, color="FFFFFF")
                cell.fill = _fill(_BEWERTUNG_FARBEN.get(bval, "9E9E9E"))
                cell.alignment = _align(halign="center", valign="center")
                dv.add(cell)
            else:
                cell.font = _font(size=9)
                cell.fill = _fill(row_bg)

        current_row += 1

    ws.freeze_panes = f"A{header_row + 1}"
    ws.print_title_rows = f"1:{header_row}"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = "landscape"

    out_path = safe_generated_file(out_path, workspace_root_from(Path(__file__)))
    ensure_private_dir(out_path.parent)
    wb.save(str(out_path))
    ensure_private_file(out_path)
    audit_event("export.write", module="nis2", outcome="success", details={"path": str(out_path), "kind": "xlsx"})
    return out_path


def import_fragebogen(xlsx_path: Path) -> list[dict[str, Any]]:
    xlsx_path = Path(xlsx_path)
    from security_utils import validate_office_archive

    validate_office_archive(xlsx_path, expected_suffix=".xlsx")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=20):
        if str(row[0].value or "").strip().upper() == "ID":
            header_row = row[0].row
            break

    if header_row is None:
        raise ValueError("Header-Zeile mit 'ID' nicht gefunden – ist dies ein NIS2-Fragebogen?")

    results: list[dict[str, Any]] = []
    known_ids = {req["id"] for req in NIS2_ANFORDERUNGEN}

    for row in ws.iter_rows(min_row=header_row + 1):
        id_val = str(row[0].value or "").strip()
        if not id_val or id_val not in known_ids:
            continue

        def _cell(col: int) -> str:
            return str(row[col].value or "").strip() if col < len(row) else ""

        try:
            bew_raw = _cell(5)
            bewertung = int(float(bew_raw)) if bew_raw else 0
            if bewertung < 0 or bewertung > 5:
                bewertung = 0
        except (ValueError, TypeError):
            bewertung = 0

        results.append({
            "anforderung_id":  id_val,
            "bewertung":       bewertung,
            "kommentar":       _cell(6),
            "massnahme":       _cell(7),
            "verantwortlich":  _cell(8),
            "zieldatum":       _cell(9),
        })

    return results
