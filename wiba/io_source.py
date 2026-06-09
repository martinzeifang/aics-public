"""WiBA-Katalogquelle: BSI-Download + Parser (WiBA-Tool-XLSX + Checklisten-ZIP).

Die BSI-Originaldateien werden **nicht** ins Repo eingecheckt, sondern zur
Laufzeit in ein Datenverzeichnis geladen (updatefähig, #1119) und in den
DB-Katalog (``wiba_themen`` + ``wiba_prueffragen``) überführt.

Quellen (BSI IT-Grundschutz / WiBA):
- WiBA-Tool (XLSX), Sheet „Dokumentation WiBA": Thema | Nr | Prüffrage | Hilfsmittel | Aufwand
- Checklisten (ZIP): je Thema eine DOCX mit Bausteinen (z. B. CON.3), Ziel, Hinweis, Links
"""
from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path
from typing import Any

BSI_BASE = "https://www.bsi.bund.de/SharedDocs/Downloads/DE/BSI/Grundschutz/WiBA"

# Logischer Name → (URL, Zieldateiname)
BSI_SOURCES: dict[str, tuple[str, str]] = {
    "tool": (f"{BSI_BASE}/WiBA_Tool_Excel.xlsx?__blob=publicationFile&v=2",
             "WiBA_Tool.xlsx"),
    "checklisten": (f"{BSI_BASE}/WiBA_Checklisten_ZIP-Datei_2023.zip?__blob=publicationFile&v=5",
                    "WiBA_Checklisten.zip"),
    "mapping": (f"{BSI_BASE}/WiBA_Mapping_2023.xlsx?__blob=publicationFile&v=3",
                "WiBA_Mapping.xlsx"),
}

CATALOG_VERSION = "WiBA 2.0 (2023)"

# #1178: Größenobergrenze für heruntergeladene Quelldateien (DoS-/Disk-Fill-Schutz).
MAX_DOWNLOAD_BYTES = 60 * 1024 * 1024


def _expected_suffix(fname: str) -> str:
    return Path(fname).suffix.casefold()


def _validate_source_archive(path: Path, fname: str) -> None:
    """#1178: ZIP/XLSX gegen Zip-Bomb/Tampering prüfen (validate_office_archive).

    Wird bei Download und vor dem Parsen aufgerufen; harte ValueError-Ablehnung.
    """
    suffix = _expected_suffix(fname)
    if suffix not in (".xlsx", ".zip"):
        return
    from security_utils import validate_office_archive
    validate_office_archive(path, expected_suffix=suffix)


def theme_key(title: str) -> str:
    """Stabiler Schlüssel je Thema (slug, tolerant ggü. '/'/Leerzeichen)."""
    return re.sub(r"[^a-z0-9äöüß]", "", str(title or "").lower())


# ── Download ──────────────────────────────────────────────────────────────────

def download_sources(dest_dir: Path, which: tuple[str, ...] = ("tool", "checklisten")
                     ) -> dict[str, Any]:
    """Lädt die BSI-WiBA-Quelldateien nach ``dest_dir``. Gibt einen Statusreport."""
    import requests  # lazy

    dest_dir = Path(dest_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)
    report: dict[str, Any] = {"ok": True, "files": [], "log": []}
    for key in which:
        if key not in BSI_SOURCES:
            continue
        url, fname = BSI_SOURCES[key]
        try:
            # #1178: gestreamt mit harter Größenobergrenze (Disk-Fill-/DoS-Schutz).
            resp = requests.get(url, timeout=60, stream=True)
            resp.raise_for_status()
            chunks: list[bytes] = []
            total = 0
            for chunk in resp.iter_content(64 * 1024):
                if not chunk:
                    continue
                total += len(chunk)
                if total > MAX_DOWNLOAD_BYTES:
                    raise ValueError(
                        f"Download überschreitet Limit ({MAX_DOWNLOAD_BYTES} B): {fname}")
                chunks.append(chunk)
            content = b"".join(chunks)
            out = dest_dir / fname
            out.write_bytes(content)
            # #1178: Archiv-Integrität/Zip-Bomb prüfen — bei Verstoß Datei verwerfen.
            try:
                _validate_source_archive(out, fname)
            except Exception:
                out.unlink(missing_ok=True)
                raise
            report["files"].append({"key": key, "file": fname, "bytes": len(content)})
            report["log"].append(f"OK {key}: {fname} ({len(content)} B)")
        except Exception as e:  # noqa: BLE001
            report["ok"] = False
            report["log"].append(f"FEHLER {key}: {type(e).__name__}: {e}")
    return report


# ── WiBA-Tool (XLSX) → Prüffragen ─────────────────────────────────────────────

def parse_wiba_tool(xlsx_path: Path) -> list[dict[str, Any]]:
    """Liest Sheet „Dokumentation WiBA" → Liste von Prüffragen.

    Erwartete Spalten: Checkliste | Nr | Prüffrage | Hilfsmittel | Aufwand | …
    """
    import openpyxl  # lazy

    # #1178: XLSX vor dem Öffnen gegen Zip-Bomb/Tampering prüfen.
    _validate_source_archive(Path(xlsx_path), Path(xlsx_path).name)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if "dokumentation" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    out: list[dict[str, Any]] = []
    order: dict[str, int] = {}
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            continue  # Header
        if not row or len(row) < 3:
            continue
        thema = (str(row[0]).strip() if row[0] is not None else "")
        nr = row[1]
        frage = (str(row[2]).strip() if row[2] is not None else "")
        hilfsmittel = (str(row[3]).strip() if len(row) > 3 and row[3] is not None else "")
        aufwand = (str(row[4]).strip() if len(row) > 4 and row[4] is not None else "")
        if not thema or not frage:
            continue
        try:
            nr_int = int(nr)
        except (TypeError, ValueError):
            continue
        tk = theme_key(thema)
        order.setdefault(tk, len(order) + 1)
        out.append({
            "control_id": f"{tk}-{nr_int}",
            "theme_key": tk,
            "theme_title": thema,
            "nr": nr_int,
            "frage": frage,
            "hilfsmittel": hilfsmittel,
            "aufwand": aufwand,
            "_order": order[tk],
        })
    return out


# ── Checklisten (ZIP/DOCX) → Thema-Metadaten ──────────────────────────────────

_HEADINGS = {
    "bausteine": "zugrundeliegende bausteine",
    "ziel": "ziel",
    "hinweis": "allgemeiner hinweis",
    "weiterfuehrend": "weiterführende informationen",
    "prueffragen": "prüffragen",
}


def _parse_checklist_docx(data: bytes, fallback_title: str) -> dict[str, Any]:
    from docx import Document  # lazy

    doc = Document(io.BytesIO(data))
    titel = fallback_title
    section: str | None = None
    bausteine: list[str] = []
    buf: dict[str, list[str]] = {"ziel": [], "hinweis": [], "weiterfuehrend": []}

    paras = list(doc.paragraphs)
    for idx, p in enumerate(paras):
        text = p.text.strip()
        if not text:
            continue
        low = text.lower()
        style = (p.style.name if p.style else "") or ""
        # Titel: Absatz nach "Checkliste:" oder Title-Style
        if "title" in style.lower() and text.lower() != "checkliste:":
            cand = text.replace("Checkliste:", "").strip()
            if cand:
                titel = cand
        # Heading-Erkennung (Formatvorlage1 / Überschrift / Heading)
        is_heading = bool(re.search(r"formatvorlage|heading|überschrift|title",
                                    style.lower()))
        matched = None
        for key, needle in _HEADINGS.items():
            if low.startswith(needle):
                matched = key
                break
        if matched:
            section = None if matched == "prueffragen" else matched
            continue
        if is_heading:
            # andere Abschnittsüberschrift → aktuellen Abschnitt beenden
            section = None
            continue
        if section == "bausteine":
            # Listenpunkte wie "CON.3 Datensicherungskonzept"
            bausteine.append(text)
        elif section in ("ziel", "hinweis", "weiterfuehrend"):
            buf[section].append(text)

    return {
        "titel": titel,
        "bausteine": "; ".join(bausteine),
        "ziel": "\n".join(buf["ziel"]).strip(),
        "hinweis": "\n".join(buf["hinweis"]).strip(),
        "weiterfuehrend": "\n".join(buf["weiterfuehrend"]).strip(),
    }


def parse_checklisten_zip(zip_path: Path) -> dict[str, dict[str, Any]]:
    """``{theme_key: {titel, bausteine, ziel, hinweis, weiterfuehrend}}``."""
    out: dict[str, dict[str, Any]] = {}
    # #1178: ZIP vor dem Entpacken gegen Zip-Bomb/Tampering prüfen.
    _validate_source_archive(Path(zip_path), Path(zip_path).name)
    with zipfile.ZipFile(zip_path) as zf:
        for name in zf.namelist():
            if not name.lower().endswith(".docx"):
                continue
            base = Path(name).stem
            fallback = re.sub(r"^Checkliste[_ ]*", "", base)
            fallback = re.sub(r"[_ ]*\d+\.\d+$", "", fallback).replace("_", " ").strip()
            meta = _parse_checklist_docx(zf.read(name), fallback)
            out[theme_key(meta["titel"])] = meta
            # zusätzlicher Key über Dateinamen (robust gegen Titel-Abweichungen)
            out.setdefault(theme_key(fallback), meta)
    return out


# ── Build Katalog ─────────────────────────────────────────────────────────────

def build_catalog(source_dir: Path) -> tuple[list[dict], list[dict]]:
    """Aus den heruntergeladenen Quelldateien → (themen, prueffragen)."""
    source_dir = Path(source_dir)
    tool = source_dir / BSI_SOURCES["tool"][1]
    zipf = source_dir / BSI_SOURCES["checklisten"][1]
    if not tool.exists():
        raise FileNotFoundError(f"WiBA-Tool nicht gefunden: {tool}")

    fragen = parse_wiba_tool(tool)
    docx_meta = parse_checklisten_zip(zipf) if zipf.exists() else {}

    themen: dict[str, dict] = {}
    for f in fragen:
        tk = f["theme_key"]
        if tk not in themen:
            md = docx_meta.get(tk, {})
            themen[tk] = {
                "theme_key": tk,
                "titel": f["theme_title"],
                "bausteine": md.get("bausteine", ""),
                "ziel": md.get("ziel", ""),
                "hinweis": md.get("hinweis", ""),
                "weiterfuehrend": md.get("weiterfuehrend", ""),
                "reihenfolge": f["_order"],
            }
    prueffragen = [{k: v for k, v in f.items() if k not in ("theme_title", "_order")}
                   for f in fragen]
    themen_list = sorted(themen.values(), key=lambda t: t["reihenfolge"])
    return themen_list, prueffragen
